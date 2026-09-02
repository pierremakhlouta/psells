"""One-time import of the PSells Excel workbook into the JSON data files.

Run once, from the project root:

    python3 import_excel.py /path/to/PSells_Inventory_CLEAN.xlsx

Reads the cleaned workbook, builds every record in memory, validates the whole
result, and only then writes data/inventory.json and data/sales.json. If any
check fails, nothing is written at all: a half-finished import is worse than no
import, because there would be no way to tell which half is real.

The workbook path is a command-line argument rather than a constant so that no
personal folder path ends up in the public repository.
"""

import os
import sys
import json
from datetime import datetime

import openpyxl


INVENTORY_FILE = "data/inventory.json"
SALES_FILE = "data/sales.json"

INVENTORY_SHEET = "Inventory"
SALES_SHEET = "Sales"

PARTNER_SHARE_MODES = ["default", "custom_percent", "custom_amount"]

# Verified against the cleaned workbook before this script was written. They are
# asserted rather than trusted so that a truncated or edited sheet cannot import
# silently.
EXPECTED_INVENTORY_RECORDS = 260
EXPECTED_SALES_RECORDS = 80
EXPECTED_SALES_UNITS = 81

MONEY_TOLERANCE = 0.01


def read_sheet(worksheet):
    headers = {}

    for column in range(1, worksheet.max_column + 1):
        header = worksheet.cell(row=1, column=column).value

        if header is not None:
            headers[str(header).strip()] = column

    rows = []

    for row in range(2, worksheet.max_row + 1):
        values = {
            header: worksheet.cell(row=row, column=column).value
            for header, column in headers.items()
        }

        if all(value in (None, "") for value in values.values()):
            continue

        values["_row"] = row
        rows.append(values)

    return rows


def require(condition, row, message):
    if not condition:
        raise ValueError(f"Row {row}: {message}")


def to_int(value, row, field):
    require(
        isinstance(value, int) and not isinstance(value, bool),
        row,
        f"{field} must be a whole number, found {value!r}"
    )

    return value


def to_float(value, row, field):
    require(
        isinstance(value, (int, float)) and not isinstance(value, bool),
        row,
        f"{field} must be a number, found {value!r}"
    )

    return float(value)


def to_text(value, row, field, allow_blank=False):
    text = "" if value is None else str(value).strip()

    if not allow_blank:
        require(text != "", row, f"{field} cannot be blank")

    return text


def to_bool(value, row, field):
    text = to_text(value, row, field).lower()

    require(
        text in ("yes", "no"),
        row,
        f"{field} must be Yes or No, found {value!r}"
    )

    return text == "yes"


def build_inventory(rows):
    inventory = []

    for index, row in enumerate(rows, start=1):
        line = row["_row"]

        discontinued = to_bool(row["Discontinued"], line, "Discontinued")

        retail_value = row["Retail (CAD)"]

        # A discontinued item has no retail price to record, so the sheet leaves
        # it blank and PSells stores zero.
        if retail_value is None or retail_value == "":
            require(
                discontinued,
                line,
                "Retail (CAD) is blank but the item is not discontinued"
            )
            retail_price = 0.0
        else:
            retail_price = to_float(retail_value, line, "Retail (CAD)")

        mode = to_text(row["Partner Share Mode"], line, "Partner Share Mode").lower()

        require(
            mode in PARTNER_SHARE_MODES,
            line,
            f"unknown partner share mode {mode!r}"
        )

        product = {
            "id": index,
            "category": to_text(row["Category"], line, "Category"),
            "name": to_text(row["Item Name"], line, "Item Name"),
            "quantity_received": to_int(
                row["Quantity Received"], line, "Quantity Received"
            ),
            "retail_price": retail_price,
            "listed_price": to_float(row["Listed Price"], line, "Listed Price"),
            "condition": to_text(row["Condition"], line, "Condition"),
            "notes": to_text(row["Notes"], line, "Notes", allow_blank=True),
            "quantity_sold": to_int(row["Quantity Sold"], line, "Quantity Sold"),
            "discontinued": discontinued,
            "partner_share_mode": mode
        }

        share_value = row["Partner Share Value"]

        if mode == "default":
            require(
                share_value is None or share_value == "",
                line,
                "default mode must not carry a partner share value"
            )
        else:
            # The key is omitted entirely for default mode, so anything reading
            # inventory must not assume it exists.
            product["partner_share_value"] = to_float(
                share_value, line, "Partner Share Value"
            )

        inventory.append(product)

    return inventory


def build_sales(rows, id_by_name):
    sales = []

    for index, row in enumerate(rows, start=1):
        line = row["_row"]

        name = to_text(row["Item Name"], line, "Item Name")

        require(
            name in id_by_name,
            line,
            f"sale refers to {name!r}, which is not in the inventory sheet"
        )

        sale_date = row["Date"]

        require(
            isinstance(sale_date, datetime),
            line,
            f"Date must be a real date, found {sale_date!r}"
        )

        sale = {
            "id": index,
            "date": sale_date.strftime("%Y-%m-%d"),
            "item_id": id_by_name[name],
            "quantity": to_int(row["Quantity Sold"], line, "Quantity Sold"),
            "sale_price": to_float(
                row["Sale Price (per unit)"], line, "Sale Price (per unit)"
            ),
            "partner_share": to_float(
                row["Partner Share (Per Unit)"], line, "Partner Share (Per Unit)"
            )
        }

        sales.append(sale)

    return sales


def validate(inventory, sales, sheet_totals):
    errors = []

    def check(condition, message):
        if not condition:
            errors.append(message)

    check(
        len(inventory) == EXPECTED_INVENTORY_RECORDS,
        f"expected {EXPECTED_INVENTORY_RECORDS} inventory records, "
        f"built {len(inventory)}"
    )

    check(
        len(sales) == EXPECTED_SALES_RECORDS,
        f"expected {EXPECTED_SALES_RECORDS} sales records, built {len(sales)}"
    )

    ids = [product["id"] for product in inventory]
    check(len(set(ids)) == len(ids), "inventory ids are not unique")
    check(ids == list(range(1, len(inventory) + 1)), "inventory ids are not 1..n")

    names = [product["name"] for product in inventory]
    check(len(set(names)) == len(names), "inventory names are not unique")

    for product in inventory:
        label = f"inventory id {product['id']}"

        check(
            product["quantity_received"] >= 1,
            f"{label}: quantity_received is below 1"
        )

        check(
            0 <= product["quantity_sold"] <= product["quantity_received"],
            f"{label}: quantity_sold is outside 0..quantity_received"
        )

        check(product["retail_price"] >= 0, f"{label}: negative retail_price")
        check(product["listed_price"] >= 0, f"{label}: negative listed_price")

        check(
            isinstance(product["discontinued"], bool),
            f"{label}: discontinued is not a boolean"
        )

        mode = product["partner_share_mode"]

        if mode == "default":
            check(
                "partner_share_value" not in product,
                f"{label}: default mode carries a partner_share_value key"
            )
        else:
            check(
                "partner_share_value" in product,
                f"{label}: {mode} is missing partner_share_value"
            )

            value = product.get("partner_share_value", 0)

            if mode == "custom_percent":
                check(
                    0 <= value <= 100,
                    f"{label}: custom_percent value outside 0..100"
                )
            else:
                check(value >= 0, f"{label}: negative custom_amount value")

        # The application refuses to put a discontinued item on any other mode,
        # so an import must not create one either.
        if product["discontinued"]:
            check(
                mode == "custom_amount",
                f"{label}: discontinued item is on {mode}"
            )

            check(
                product["retail_price"] == 0.0,
                f"{label}: discontinued item has a non-zero retail_price"
            )

    sale_ids = [sale["id"] for sale in sales]
    check(len(set(sale_ids)) == len(sale_ids), "sales ids are not unique")
    check(
        sale_ids == list(range(1, len(sales) + 1)),
        "sales ids are not 1..n"
    )

    known_ids = set(ids)

    for sale in sales:
        label = f"sale id {sale['id']}"

        check(sale["item_id"] in known_ids, f"{label}: orphaned item_id")
        check(sale["quantity"] >= 1, f"{label}: quantity below 1")
        check(sale["sale_price"] >= 0, f"{label}: negative sale_price")
        check(sale["partner_share"] >= 0, f"{label}: negative partner_share")

        try:
            datetime.strptime(sale["date"], "%Y-%m-%d")
        except ValueError:
            check(False, f"{label}: date {sale['date']!r} is not YYYY-MM-DD")

    sold_by_item = {}

    for sale in sales:
        sold_by_item[sale["item_id"]] = (
            sold_by_item.get(sale["item_id"], 0) + sale["quantity"]
        )

    for product in inventory:
        recorded = product["quantity_sold"]
        from_sales = sold_by_item.get(product["id"], 0)

        check(
            recorded == from_sales,
            f"inventory id {product['id']} ({product['name']}): "
            f"quantity_sold is {recorded} but its sales total {from_sales}"
        )

    total_units = sum(sale["quantity"] for sale in sales)

    check(
        total_units == EXPECTED_SALES_UNITS,
        f"expected {EXPECTED_SALES_UNITS} units sold, found {total_units}"
    )

    revenue = sum(sale["quantity"] * sale["sale_price"] for sale in sales)
    partner_share = sum(
        sale["quantity"] * sale["partner_share"] for sale in sales
    )

    # The sheet's own calculated columns are never imported, but they are a
    # useful independent check that the raw fields were read correctly.
    check(
        abs(revenue - sheet_totals["revenue"]) < MONEY_TOLERANCE,
        "recomputed revenue does not match the workbook's own total"
    )

    check(
        abs(partner_share - sheet_totals["partner_share"]) < MONEY_TOLERANCE,
        "recomputed partner share does not match the workbook's own total"
    )

    check(
        abs((revenue - partner_share) - sheet_totals["profit"]) < MONEY_TOLERANCE,
        "recomputed profit does not match the workbook's own total"
    )

    return errors


def summarise(inventory, sales):
    modes = {}

    for product in inventory:
        mode = product["partner_share_mode"]
        modes[mode] = modes.get(mode, 0) + 1

    discontinued = sum(1 for product in inventory if product["discontinued"])

    received = sum(product["quantity_received"] for product in inventory)
    sold = sum(product["quantity_sold"] for product in inventory)

    revenue = sum(sale["quantity"] * sale["sale_price"] for sale in sales)
    partner_share = sum(
        sale["quantity"] * sale["partner_share"] for sale in sales
    )

    print()
    print("Import summary")
    print()
    print(f"Inventory records: {len(inventory)}")

    for mode in PARTNER_SHARE_MODES:
        print(f"  {mode}: {modes.get(mode, 0)}")

    print(f"  discontinued: {discontinued}")
    print()
    print(f"Sales records: {len(sales)}")
    print(f"Total received: {received}")
    print(f"Total sold: {sold}")
    print(f"Total available: {received - sold}")
    print()
    print(f"Total revenue: ${revenue:.2f}")
    print(f"Total partner share earned: ${partner_share:.2f}")
    print(f"Total profit: ${revenue - partner_share:.2f}")
    print()
    print("Payments are not imported. Enter them in PSells to see balance owing.")


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 import_excel.py <workbook.xlsx>")
        return 1

    workbook_path = sys.argv[1]

    if not os.path.exists(workbook_path):
        print(f"Workbook not found: {workbook_path}")
        return 1

    for path in (INVENTORY_FILE, SALES_FILE):
        if os.path.exists(path):
            print(f"{path} already exists. Refusing to overwrite it.")
            print("Move or delete the existing data files first.")
            return 1

    # data_only=True returns the last value Excel calculated. Most of the sales
    # partner-share cells are formulas, so without this the formula text would
    # be imported instead of the amount.
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)

    inventory_rows = read_sheet(workbook[INVENTORY_SHEET])
    sales_rows = read_sheet(workbook[SALES_SHEET])

    try:
        inventory = build_inventory(inventory_rows)
        id_by_name = {
            product["name"]: product["id"]
            for product in inventory
        }
        sales = build_sales(sales_rows, id_by_name)

    except ValueError as error:
        print(f"Could not read the workbook: {error}")
        return 1

    sheet_totals = {
        "revenue": sum(row["Total Revenue"] or 0 for row in sales_rows),
        "partner_share": sum(
            row["Total Partner Share"] or 0 for row in sales_rows
        ),
        "profit": sum(row["Profit"] or 0 for row in sales_rows)
    }

    errors = validate(inventory, sales, sheet_totals)

    if errors:
        print(f"Import aborted. {len(errors)} problem(s) found, nothing written.")
        print()

        for message in errors:
            print(f"  {message}")

        return 1

    os.makedirs(os.path.dirname(INVENTORY_FILE), exist_ok=True)

    save_data(inventory, INVENTORY_FILE)
    save_data(sales, SALES_FILE)

    print(f"Wrote {INVENTORY_FILE} and {SALES_FILE}.")

    summarise(inventory, sales)

    return 0


def save_data(data, filepath):
    with open(filepath, "w") as f:
        json.dump(data, f, indent=2)


if __name__ == "__main__":
    sys.exit(main())
